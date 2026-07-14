from pathlib import Path

from src.tools.file_readers.base_reader import BaseFileReader

class ExcelReader(BaseFileReader):
    @property
    def name(self) -> str:
        return "excel_reader"

    @property
    def description(self) -> str:
        return "Extract sheet names and data summary from Excel files."

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    def read(self, file_path: str, max_rows: int = 50) -> str:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required to read Excel files. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sections = [f"Excel File: {path.name}", f"Sheets: {', '.join(wb.sheetnames)}", ""]

        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    rows.append([f"... ({ws.max_row - max_rows} more rows)"])
                    break
                rows.append([str(c) if c is not None else "" for c in row])

            sections.append(f"--- Sheet: {name} ---")
            for row in rows:
                sections.append("\t".join(row))
            sections.append("")

        wb.close()
        return "\n".join(sections)
